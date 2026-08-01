# ==============================================================================
# ARCHIVO: reporte_calc.py
# DESCRIPCIÓN: Generación de hoja de datos Excel/Calc compatible con nombres
#              de claves uniformizadas con corchetes [...].
# ==============================================================================

import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generar_calc_hoja_datos(res_dict, meta_dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "TEMA_Data_Sheet"

    font_titulo = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    font_seccion = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_etiqueta = Font(name="Arial", size=9, bold=True, color="1A202C")
    font_valor = Font(name="Arial", size=9, bold=False, color="1A202C")
    
    fill_header = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
    fill_seccion = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    fill_celda_alt = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'), right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'), bottom=Side(style='thin', color='CBD5E0')
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = "TEMA EQUIPMENT DATA SHEET - INTERCAMBIADOR DE CASCO Y TUBOS"
    ws["A1"].font = font_titulo
    ws["A1"].fill = fill_header
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta_filas = [
        ("TAG DEL EQUIPO [-]:", meta_dict.get("tag", "HEX-0100"), "FECHA DE EMISIÓN [-]:", datetime.date.today().strftime("%d/%m/%Y")),
        ("PROYECTO / PLANTA [-]:", meta_dict.get("proyecto", "PROYECTO GENERAL"), "NORMAS APLICADAS [-]:", "TEMA Class R / ASME VIII"),
        ("CALCULADO POR [-]:", meta_dict.get("calculado_por", "E. Livingston"), "REVISADO POR [-]:", meta_dict.get("revisado_por", "")),
        ("APROBADO POR [-]:", meta_dict.get("aprobado_por", ""), "ESTADO / REVISIÓN [-]:", meta_dict.get("revision", "0 - COTIZACIÓN"))
    ]

    row_idx = 2
    for k1, v1, k2, v2 in meta_filas:
        ws.cell(row=row_idx, column=1, value=k1).font = font_etiqueta
        ws.cell(row=row_idx, column=2, value=v1).font = font_valor
        ws.cell(row=row_idx, column=3, value=k2).font = font_etiqueta
        ws.cell(row=row_idx, column=4, value=v2).font = font_valor
        for c in range(1, 5):
            ws.cell(row=row_idx, column=c).border = thin_border
            ws.cell(row=row_idx, column=c).fill = fill_celda_alt
        row_idx += 1

    def agregar_seccion(titulo, filas, r_idx):
        r_idx += 1
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
        celda_tit = ws.cell(row=r_idx, column=1, value=titulo.upper())
        celda_tit.font = font_seccion
        celda_tit.fill = fill_seccion
        celda_tit.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r_idx].height = 20
        r_idx += 1

        for k1, v1, k2, v2 in filas:
            ws.cell(row=r_idx, column=1, value=k1).font = font_etiqueta
            ws.cell(row=r_idx, column=2, value=v1).font = font_valor
            ws.cell(row=r_idx, column=3, value=k2).font = font_etiqueta
            ws.cell(row=r_idx, column=4, value=v2).font = font_valor
            for c in range(1, 5):
                ws.cell(row=r_idx, column=c).border = thin_border
            r_idx += 1
        return r_idx

    # 1. TERMODINÁMICA
    t = res_dict["Termodinámica"]
    filas_termo = [
        ("Carga Térmica Total [kW]", t["Carga Térmica Total [kW]"], "LMTD Contracorriente [°C]", t["LMTD Contracorriente [°C]"]),
        ("Caudal Proceso Lado Caliente [kg/s]", t["Caudal Fluido Proceso [kg/s]"], "Caudal Servicio Auxiliar [kg/s]", t["Caudal Servicio Auxiliar [kg/s]"]),
        ("Temp. Entrada Caliente [°C]", t["Temp. Entrada Caliente [°C]"], "Temp. Salida Caliente [°C]", t["Temp. Salida Caliente [°C]"]),
        ("Temp. Entrada Frío [°C]", t["Temp. Entrada Frío [°C]"], "Temp. Salida Frío [°C]", t["Temp. Salida Frío [°C]"]),
        ("Factor de Corrección Ft [-]", t["Factor Corrección Ft [-]"], "LMTD Efectiva Real [°C]", t["LMTD Efectiva [°C]"])
    ]
    row_idx = agregar_seccion("1. Condiciones Operativas y Balance Térmico (CoolProp / IF97)", filas_termo, row_idx)

    # 2. GEOMETRÍA TEMA & KERN
    d = res_dict["Dimensionamiento TEMA & Kern"]
    filas_tema = [
        ("Clasificación Normativa TEMA [-]", d["Clasificación TEMA [-]"], "Pasos en Tubos [uds]", d["Pasos por Tubo [uds]"]),
        ("Área Requerida Teórica [m²]", d["Área Requerida Teórica [m²]"], "Área Instalada Real [m²]", d["Área Instalada Real [m²]"]),
        ("Número Total de Tubos [uds]", d["Número Total de Tubos [uds]"], "Diámetro Exterior Tubo OD [mm]", d["Diámetro Ext. Tubo OD [mm]"]),
        ("Diámetro Interno Casco Ds [mm]", d["Diámetro Interno Casco Ds [mm]"], "Longitud Nominal Tubos [m]", d["Longitud Nominal Tubo [m]"]),
        ("Espaciado de Bafles B [mm]", d["Espaciado de Bafles B [mm]"], "Número de Bafles [uds]", d["Número de Bafles [uds]"])
    ]
    row_idx = agregar_seccion("2. Dimensionamiento Geométrico (Sinnott Cap. 12 / Método de Kern)", filas_tema, row_idx)

    # 3. VERIFICACIÓN CONVECTIVA
    r = res_dict["Verificación Convectiva (Rating Kern)"]
    filas_rating = [
        ("Coef. Estimado de Prueba U_trial [W/m²·K]", r["Coef. Global Estimado U_trial [W/m²·K]"], "Coef. Convectivo Casco h_o [W/m²·K]", r["Coef. Convectivo Casco h_o [W/m²·K]"]),
        ("Coef. Convectivo Tubos h_i [W/m²·K]", r["Coef. Convectivo Tubos h_i [W/m²·K]"], "Coef. Global REAL U_calc [W/m²·K]", r["Coef. Global REAL U_calc [W/m²·K]"]),
        ("Margen Seguridad Térmica [%]", f'{r["Margen Seguridad Térmica [%]"]} %', "Ensuciamiento Normado Rf [m²·K/W]", r["Resistencia Ensuciamiento Rf [m²·K/W]"])
    ]
    row_idx = agregar_seccion("3. Verificación Convectiva y Sizing vs. Rating (Sinnott Cap. 12.9)", filas_rating, row_idx)

    # 4. DISEÑO MECÁNICO ASME BPVC VIII & MATERIALES
    m = res_dict["Diseño Mecánico ASME BPVC"]
    filas_mech = [
        ("Código Diseño Mecánico [-]", "ASME BPVC Sec. VIII Div. 1", "Presión Diseño ASME [bar]", m["Presión Diseño ASME [bar]"]),
        ("Material Carcasa / Casco [-]", m["Material Carcasa [-]"], "Material de los Tubos [-]", m["Material Tubos [-]"]),
        ("Tensión Adm. Carcasa S [MPa]", m["Tensión Adm. Carcasa S [MPa]"], "Tensión Adm. Tubos S [MPa]", m["Tensión Adm. Tubos S [MPa]"]),
        ("Conductividad Tubo k [W/m·K]", m["Conductividad Tubo k [W/m·K]"], "Sobreespesor Corrosión [mm]", m["Sobreespesor Corrosión [mm]"]),
        ("Espesor Comercial Casco [mm]", m["Espesor Casco Comercial [mm]"], "Espesor Tubos BWG [mm]", m["Espesor Tubo Comercial BWG [mm]"]),
        ("Inversión Estimada CAPEX [USD]", f"${m['CAPEX Estimado [USD]']:,.2f}", "Metodología CAPEX [-]", "Sinnott Cap. 6 (Class 4/5)")
    ]
    row_idx = agregar_seccion("4. Diseño Mecánico, Materiales y Presupuesto Económico (ASME VIII / Sinnott Cap. 6)", filas_mech, row_idx)

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer